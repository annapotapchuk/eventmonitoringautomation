import asyncio
import logging
import os
from openpyxl import Workbook, load_workbook
from scraper.gefma import scrape_gefma_events
from scraper.ifma import scrape_ifma_events
from scraper.realfm import scrape_realfm_events
from scraper.iwfm import scrape_iwfm_events
from scraper.eurofm import scrape_eurofm_events
from scraper.fmj import scrape_fmj_events
from scraper.fmuk import scrape_fmuk_events
from scraper.facility_manager import scrape_facility_manager_events
from scraper.ifmnet import scrape_ifmnet_events
from scraper.builtworld import scrape_builtworld_events
from scraper.i2fm import scrape_i2fm_events

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

OUTPUT_FILE = 'events.xlsx'

def save_to_excel(events: list, filepath: str):
    """
    Saves events to an Excel file, updating existing file if present.
    Deduplicates based on (title, url, source) composite key.
    """
    # Load existing events if file exists
    existing_events = {}
    if os.path.exists(filepath):
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            # Skip header row, read existing events
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:  # Skip empty rows
                    title, date, location, url, source = row
                    key = (title, url, source)
                    existing_events[key] = {
                        'title': title,
                        'date': date,
                        'location': location,
                        'url': url,
                        'source': source
                    }
            logger.info(f"Loaded {len(existing_events)} existing events from {filepath}")
        except Exception as e:
            logger.warning(f"Could not read existing file, creating new: {e}")
    
    # Add new events (update existing if same key but new data)
    new_count = 0
    updated_count = 0
    for event in events:
        key = (event['title'], event['url'], event['source'])
        if key not in existing_events:
            existing_events[key] = event
            new_count += 1
        else:
            # Update if we now have better data (replace "See details" with actual data)
            existing = existing_events[key]
            updated = False
            if existing.get('date') == 'See details' and event.get('date') != 'See details':
                existing['date'] = event['date']
                updated = True
            if existing.get('location') == 'See details' and event.get('location') != 'See details':
                existing['location'] = event['location']
                updated = True
            if updated:
                updated_count += 1
    
    logger.info(f"Added {new_count} new events, updated {updated_count} existing events")
    
    # Sort events by date (closest first, unparseable dates at end)
    from datetime import datetime
    import re
    
    def parse_date(date_str: str):
        """Parse various date formats and return a sortable date or None."""
        if not date_str or date_str == 'See details':
            return None
        
        # Try DD.MM.YYYY format (GEFMA)
        match = re.match(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', date_str)
        if match:
            try:
                return datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))
            except ValueError:
                pass
        
        # Try "DD./DD. Month YYYY" format (e.g., "11./12. März 2026")
        match = re.match(r'(\d{1,2})\.?/?\.?\s*(\d{1,2})?\.?\s*(\w+)\s+(\d{4})', date_str)
        if match:
            day = int(match.group(1))
            month_name = match.group(3).lower()
            year = int(match.group(4))
            month_map = {
                'januar': 1, 'january': 1, 'jan': 1,
                'februar': 2, 'february': 2, 'feb': 2,
                'märz': 3, 'march': 3, 'mar': 3,
                'april': 4, 'apr': 4,
                'mai': 5, 'may': 5,
                'juni': 6, 'june': 6, 'jun': 6,
                'juli': 7, 'july': 7, 'jul': 7,
                'august': 8, 'aug': 8,
                'september': 9, 'sep': 9, 'sept': 9,
                'oktober': 10, 'october': 10, 'oct': 10,
                'november': 11, 'nov': 11,
                'dezember': 12, 'december': 12, 'dec': 12
            }
            if month_name in month_map:
                try:
                    return datetime(year, month_map[month_name], day)
                except ValueError:
                    pass
        
        return None
    
    def sort_key(event):
        parsed = parse_date(event.get('date', ''))
        if parsed is None:
            return (1, datetime.max)  # Put unparseable dates at the end
        return (0, parsed)
    
    sorted_events = sorted(existing_events.values(), key=sort_key)
    logger.info(f"Sorted {len(sorted_events)} events by date (closest first)")
    
    # Write all events to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"
    
    # Write header
    headers = ['Title', 'Date', 'Location', 'URL', 'Source']
    ws.append(headers)
    
    # Style header row
    from openpyxl.styles import Font
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    
    # Write sorted data
    for event in sorted_events:
        ws.append([
            event.get('title', ''),
            event.get('date', ''),
            event.get('location', ''),
            event.get('url', ''),
            event.get('source', '')
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50  # Title
    ws.column_dimensions['B'].width = 20  # Date
    ws.column_dimensions['C'].width = 30  # Location
    ws.column_dimensions['D'].width = 60  # URL
    ws.column_dimensions['E'].width = 10  # Source
    
    wb.save(filepath)
    logger.info(f"Saved {len(sorted_events)} total events to {filepath}")

async def main():
    logger.info("Starting scraper run...")
    
    all_events = []
    
    # Run GEFMA
    try:
        gefma_events = scrape_gefma_events()
        all_events.extend(gefma_events)
        logger.info(f"Retrieved {len(gefma_events)} events from GEFMA.")
    except Exception as e:
        logger.error(f"Failed to scrape GEFMA: {e}")
        
    # Run IFMA
    try:
        ifma_events = scrape_ifma_events()
        all_events.extend(ifma_events)
        logger.info(f"Retrieved {len(ifma_events)} events from IFMA.")
    except Exception as e:
        logger.error(f"Failed to scrape IFMA: {e}")
    
    # Run RealFM
    try:
        realfm_events = scrape_realfm_events()
        all_events.extend(realfm_events)
        logger.info(f"Retrieved {len(realfm_events)} events from RealFM.")
    except Exception as e:
        logger.error(f"Failed to scrape RealFM: {e}")
    
    # Run IWFM
    try:
        iwfm_events = scrape_iwfm_events()
        all_events.extend(iwfm_events)
        logger.info(f"Retrieved {len(iwfm_events)} events from IWFM.")
    except Exception as e:
        logger.error(f"Failed to scrape IWFM: {e}")
    
    # Run EuroFM
    try:
        eurofm_events = scrape_eurofm_events()
        all_events.extend(eurofm_events)
        logger.info(f"Retrieved {len(eurofm_events)} events from EuroFM.")
    except Exception as e:
        logger.error(f"Failed to scrape EuroFM: {e}")
    
    # Run FMJ
    try:
        fmj_events = scrape_fmj_events()
        all_events.extend(fmj_events)
        logger.info(f"Retrieved {len(fmj_events)} events from FMJ.")
    except Exception as e:
        logger.error(f"Failed to scrape FMJ: {e}")
    
    # Run FMUK
    try:
        fmuk_events = scrape_fmuk_events()
        all_events.extend(fmuk_events)
        logger.info(f"Retrieved {len(fmuk_events)} events from FMUK.")
    except Exception as e:
        logger.error(f"Failed to scrape FMUK: {e}")
    
    # Run Facility-Manager.de
    try:
        fm_de_events = scrape_facility_manager_events()
        all_events.extend(fm_de_events)
        logger.info(f"Retrieved {len(fm_de_events)} events from Facility-Manager.de.")
    except Exception as e:
        logger.error(f"Failed to scrape Facility-Manager.de: {e}")
    
    # Run i-FM.net
    try:
        ifmnet_events = scrape_ifmnet_events()
        all_events.extend(ifmnet_events)
        logger.info(f"Retrieved {len(ifmnet_events)} events from i-FM.net.")
    except Exception as e:
        logger.error(f"Failed to scrape i-FM.net: {e}")
    
    # Run Builtworld (uses Selenium)
    try:
        builtworld_events = scrape_builtworld_events()
        all_events.extend(builtworld_events)
        logger.info(f"Retrieved {len(builtworld_events)} events from Builtworld.")
    except Exception as e:
        logger.error(f"Failed to scrape Builtworld: {e}")
    
    # Run i2FM
    try:
        i2fm_events = scrape_i2fm_events()
        all_events.extend(i2fm_events)
        logger.info(f"Retrieved {len(i2fm_events)} events from i2FM.")
    except Exception as e:
        logger.error(f"Failed to scrape i2FM: {e}")
        
    # Save to Excel
    save_to_excel(all_events, OUTPUT_FILE)
    logger.info(f"Scraping complete. Output saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    asyncio.run(main())
